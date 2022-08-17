from administration.models import Afgiftsperiode, ProduktType, SkemaType
from administration.models import BeregningsModel2021
from administration.models import G69Code
from administration.models import G69CodeExport
from datetime import date
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.test import TestCase
from django.urls import reverse
from indberetning.models import Indhandlingssted
from indberetning.models import Virksomhed, Indberetning, IndberetningLinje
from io import BytesIO
from openpyxl import load_workbook


class G69TestCase(TestCase):
    def setUp(self):
        super().setUp()
        self.user = get_user_model().objects.create(username="TestUser")
        self.skematyper = {
            1: SkemaType.objects.get(id=1),
            2: SkemaType.objects.get(id=2),
            3: SkemaType.objects.get(id=3),
        }
        self.sted = Indhandlingssted.objects.get(navn="Innaarsuit", stedkode=10769)
        self.virksomhed, _ = Virksomhed.objects.get_or_create(cvr=1234, sted=self.sted)
        self.beregningsmodel, _ = BeregningsModel2021.objects.get_or_create(
            navn="TestBeregningsModel"
        )
        self.periode = Afgiftsperiode.objects.create(
            beregningsmodel=self.beregningsmodel,
            navn_dk="testperiode",
            dato_fra=date(2022, 1, 1),
            dato_til=date(2022, 3, 31),
        )


class FangsttypeTestCase(G69TestCase):
    def test_fangsttyper(self):

        indberetning1 = Indberetning.objects.create(
            afgiftsperiode=self.periode,
            skematype=self.skematyper[1],
            virksomhed=self.virksomhed,
        )
        for navn in [
            "Makrel",
            "Sild",
            "Lodde",
            "Blåhvilling",
            "Guldlaks",
            "Hellefisk",
            "Torsk",
            "Kuller",
            "Sej",
            "Rødfisk",
            "Reje - havgående licens",
        ]:
            for produkttype in ProduktType.objects.filter(navn_dk__startswith=navn):
                linje = IndberetningLinje.objects.create(
                    indberetning=indberetning1,
                    produkttype=produkttype,
                    levende_vægt=1000,
                    salgspris=10000,
                )
                self.assertEqual(linje.fangsttype, "havgående", produkttype.navn_dk)

        indberetning2 = Indberetning.objects.create(
            afgiftsperiode=self.periode,
            skematype=self.skematyper[2],
            virksomhed=self.virksomhed,
        )
        for navn in [
            "Makrel",
            "Sild",
            "Lodde",
            "Blåhvilling",
            "Guldlaks",
            "Hellefisk",
            "Torsk",
            "Kuller",
            "Sej",
            "Rødfisk",
        ]:
            for produkttype in ProduktType.objects.filter(navn_dk__startswith=navn):
                linje = IndberetningLinje.objects.create(
                    indberetning=indberetning2,
                    produkttype=produkttype,
                    levende_vægt=1000,
                    salgspris=10000,
                )
                self.assertEqual(linje.fangsttype, "indhandling", produkttype.navn_dk)

        indberetning3 = Indberetning.objects.create(
            afgiftsperiode=self.periode,
            skematype=self.skematyper[3],
            virksomhed=self.virksomhed,
        )
        for navn in [
            "Makrel",
            "Sild",
            "Lodde",
            "Blåhvilling",
            "Guldlaks",
            "Hellefisk",
            "Torsk",
            "Kuller",
            "Sej",
            "Rødfisk",
            "Reje - kystnær licens",
        ]:
            for produkttype in ProduktType.objects.filter(fiskeart__navn_dk=navn):
                linje = IndberetningLinje.objects.create(
                    indberetning=indberetning3,
                    produkttype=produkttype,
                    levende_vægt=1000,
                    salgspris=10000,
                )
                self.assertEqual(linje.fangsttype, "kystnært", produkttype.navn_dk)

        indberetning4 = Indberetning.objects.create(
            afgiftsperiode=self.periode,
            skematype=self.skematyper[1],
            virksomhed=self.virksomhed,
        )
        navn = "Reje - Svalbard og Barentshavet"
        for produkttype in ProduktType.objects.filter(navn_dk__startswith=navn):
            linje = IndberetningLinje.objects.create(
                indberetning=indberetning4,
                produkttype=produkttype,
                levende_vægt=1000,
                salgspris=10000,
            )
            self.assertEqual(linje.fangsttype, "svalbard", produkttype.navn_dk)


class G69KodeTestCase(G69TestCase):
    def setUp(self):
        super().setUp()
        self.username = "test"
        self.user = get_user_model().objects.create_user(username=self.username)
        self.password = "test"
        self.user.set_password(self.password)
        administration_group, _ = Group.objects.get_or_create(name="administration")
        self.user.save()
        self.user.groups.add(administration_group)

    def test_get_kode(self):
        skematype_havgående = 1
        skematype_indhandling = 2
        skematype_kystnær = 3

        for navn, grønlandsk, skematype_id, expected_kode in (
            ("Hellefisk", None, skematype_havgående, 10011),
            ("Hellefisk", None, skematype_indhandling, 10013),
            ("Hellefisk", None, skematype_kystnær, 10012),
            ("Makrel", True, skematype_havgående, 10021),
            ("Makrel", True, skematype_indhandling, 10021),
            ("Makrel", False, skematype_havgående, 10022),
            ("Makrel", False, skematype_indhandling, 10022),
            ("Lodde", True, skematype_havgående, 10025),
            ("Lodde", True, skematype_indhandling, 10025),
            ("Lodde", False, skematype_havgående, 10026),
            ("Lodde", False, skematype_indhandling, 10026),
            ("Torsk", None, skematype_havgående, 10027),
            ("Torsk", None, skematype_indhandling, 10028),
            ("Rødfisk", None, skematype_havgående, 10029),
            ("Rødfisk", None, skematype_indhandling, 10030),
            ("Kuller", None, skematype_havgående, 10031),
            ("Kuller", None, skematype_indhandling, 10032),
            ("Sej", None, skematype_havgående, 10033),
            ("Sej", None, skematype_indhandling, 10034),
            ("Blåhvilling", True, skematype_havgående, 10038),
            ("Blåhvilling", True, skematype_indhandling, 10038),
            ("Blåhvilling", False, skematype_havgående, 10039),
            ("Blåhvilling", False, skematype_indhandling, 10039),
            ("Guldlaks", True, skematype_havgående, 10040),
            ("Guldlaks", True, skematype_indhandling, 10040),
            ("Guldlaks", False, skematype_havgående, 10041),
            ("Guldlaks", False, skematype_indhandling, 10041),
        ):
            produkttype = ProduktType.objects.filter(
                fiskeart__navn_dk=navn, fartoej_groenlandsk=grønlandsk
            ).first()
            expected_full_kode = (
                str(self.periode.dato_fra.year)[2:]
                + str(self.sted.stedkode)[-6:].zfill(6)
                + str(produkttype.fiskeart.kode).zfill(2)
                + str(expected_kode).zfill(5)
            )
            linje = IndberetningLinje.objects.create(
                indberetning=Indberetning.objects.create(
                    afgiftsperiode=self.periode,
                    skematype=self.skematyper[skematype_id],
                    virksomhed=self.virksomhed,
                ),
                produkttype=produkttype,
                levende_vægt=1000,
                salgspris=10000,
                indhandlingssted=self.sted,
            )
            self.assertEqual(
                G69Code.for_indberetningslinje(linje).kode, expected_full_kode, navn
            )

    def test_get_raw(self):
        expected_headers = [
            ("kode", "G69-Kode"),
            ("skatteår", "År"),
            ("fangsttype", "Fangsttype"),
            ("aktivitet_kode", "Aktivitetskode"),
            ("fiskeart_navn", "Fiskeart"),
            ("fiskeart_kode", "Fiskeart kode"),
            ("sted_navn", "Sted"),
            ("sted_kode", "Stedkode"),
            ("grønlandsk", "Grønlandsk fartøj"),
        ]
        expected_data_subset = [
            # Stedkoder ('kode') er skrevet på denne måde for at øge læsbarheden; den er sammensat af de fire viste komponenter
            {
                "kode": "22" + "010769" + "01" + "10021",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010021",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "01" + "10021",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010021",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "01" + "10021",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010021",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "01" + "10022",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010022",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "01" + "10022",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010022",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "01" + "10022",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010022",
                "fiskeart_navn": "Makrel",
                "fiskeart_kode": 1,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "02" + "10023",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010023",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "02" + "10023",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010023",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "02" + "10023",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010023",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "02" + "10024",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010024",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "02" + "10024",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010024",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "02" + "10024",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010024",
                "fiskeart_navn": "Sild",
                "fiskeart_kode": 2,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "03" + "10025",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010025",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "03" + "10025",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010025",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "03" + "10025",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010025",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "03" + "10026",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010026",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "03" + "10026",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010026",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "03" + "10026",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010026",
                "fiskeart_navn": "Lodde",
                "fiskeart_kode": 3,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "04" + "10038",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010038",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "04" + "10038",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010038",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "04" + "10038",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010038",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "04" + "10039",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010039",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "04" + "10039",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010039",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "04" + "10039",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010039",
                "fiskeart_navn": "Blåhvilling",
                "fiskeart_kode": 4,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "05" + "10040",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010040",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "05" + "10040",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010040",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "05" + "10040",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010040",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Ja",
            },
            {
                "kode": "22" + "010769" + "05" + "10041",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010041",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "05" + "10041",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010041",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "05" + "10041",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010041",
                "fiskeart_navn": "Guldlaks",
                "fiskeart_kode": 5,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "Nej",
            },
            {
                "kode": "22" + "010769" + "06" + "10011",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010011",
                "fiskeart_navn": "Hellefisk",
                "fiskeart_kode": 6,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "06" + "10012",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010012",
                "fiskeart_navn": "Hellefisk",
                "fiskeart_kode": 6,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "06" + "10013",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010013",
                "fiskeart_navn": "Hellefisk",
                "fiskeart_kode": 6,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "07" + "10027",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010027",
                "fiskeart_navn": "Torsk",
                "fiskeart_kode": 7,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "07" + "10028",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010028",
                "fiskeart_navn": "Torsk",
                "fiskeart_kode": 7,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "08" + "10031",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010031",
                "fiskeart_navn": "Kuller",
                "fiskeart_kode": 8,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "08" + "10032",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010032",
                "fiskeart_navn": "Kuller",
                "fiskeart_kode": 8,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "09" + "10033",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010033",
                "fiskeart_navn": "Sej",
                "fiskeart_kode": 9,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "09" + "10034",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010034",
                "fiskeart_navn": "Sej",
                "fiskeart_kode": 9,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "10" + "10029",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010029",
                "fiskeart_navn": "Rødfisk",
                "fiskeart_kode": 10,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "10" + "10030",
                "skatteår": 2022,
                "fangsttype": "indhandling",
                "aktivitet_kode": "010030",
                "fiskeart_navn": "Rødfisk",
                "fiskeart_kode": 10,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "11" + "10011",
                "skatteår": 2022,
                "fangsttype": "havgående",
                "aktivitet_kode": "010011",
                "fiskeart_navn": "Reje - havgående licens",
                "fiskeart_kode": 11,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "12" + "10012",
                "skatteår": 2022,
                "fangsttype": "kystnært",
                "aktivitet_kode": "010012",
                "fiskeart_navn": "Reje - kystnær licens",
                "fiskeart_kode": 12,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
            {
                "kode": "22" + "010769" + "13" + "10014",
                "skatteår": 2022,
                "fangsttype": "svalbard",
                "aktivitet_kode": "010014",
                "fiskeart_navn": "Reje - Svalbard og Barentshavet",
                "fiskeart_kode": 13,
                "sted_navn": "Innaarsuit",
                "sted_kode": 10769,
                "grønlandsk": "",
            },
        ]

        self.maxDiff = None
        raw_data = G69Code.get_spreadsheet_raw(2022)
        self.assertListEqual(raw_data["headers"], expected_headers)
        for item in expected_data_subset:
            self.assertIn(item, raw_data["data"])

    def test_csv(self):
        self.client.login(username=self.username, password=self.password)
        self.client.post(reverse("administration:g69-create"), {"år": 2022})
        pk = G69CodeExport.objects.filter(år=2022).first().pk

        response = self.client.get(
            reverse("administration:g69-download", kwargs={"pk": pk})
        )
        self.assertEqual(
            response.get("Content-Disposition"), 'attachment; filename="g69_koder.xlsx"'
        )

        wb = load_workbook(filename=BytesIO(response.getvalue()))
        ws = wb.active
        self.assertEqual(ws.max_column, 9)
        self.assertEqual(ws.max_row, 2665)

        # Koder er unikke og sorterede
        expected = [
            [
                "G69-Kode",
                "År",
                "Fangsttype",
                "Aktivitetskode",
                "Fiskeart",
                "Fiskeart kode",
                "Sted",
                "Stedkode",
                "Grønlandsk fartøj",
            ],
            [
                "22" + "010769" + "01" + "10021",
                "2022",
                None,
                "010021",
                "Makrel",
                1,
                "Innaarsuit",
                10769,
                "Ja",
            ],
            [
                "22" + "010769" + "01" + "10022",
                "2022",
                None,
                "010022",
                "Makrel",
                1,
                "Innaarsuit",
                10769,
                "Nej",
            ],
            [
                "22" + "010769" + "02" + "10023",
                "2022",
                None,
                "010023",
                "Sild",
                2,
                "Innaarsuit",
                10769,
                "Ja",
            ],
            [
                "22" + "010769" + "02" + "10024",
                "2022",
                None,
                "010024",
                "Sild",
                2,
                "Innaarsuit",
                10769,
                "Nej",
            ],
            [
                "22" + "010769" + "03" + "10025",
                "2022",
                None,
                "010025",
                "Lodde",
                3,
                "Innaarsuit",
                10769,
                "Ja",
            ],
            [
                "22" + "010769" + "03" + "10026",
                "2022",
                None,
                "010026",
                "Lodde",
                3,
                "Innaarsuit",
                10769,
                "Nej",
            ],
            [
                "22" + "010769" + "04" + "10038",
                "2022",
                None,
                "010038",
                "Blåhvilling",
                4,
                "Innaarsuit",
                10769,
                "Ja",
            ],
            [
                "22" + "010769" + "04" + "10039",
                "2022",
                None,
                "010039",
                "Blåhvilling",
                4,
                "Innaarsuit",
                10769,
                "Nej",
            ],
            [
                "22" + "010769" + "05" + "10040",
                "2022",
                None,
                "010040",
                "Guldlaks",
                5,
                "Innaarsuit",
                10769,
                "Ja",
            ],
            [
                "22" + "010769" + "05" + "10041",
                "2022",
                None,
                "010041",
                "Guldlaks",
                5,
                "Innaarsuit",
                10769,
                "Nej",
            ],
            [
                "22" + "010769" + "06" + "10011",
                "2022",
                "havgående",
                "010011",
                "Hellefisk",
                6,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "06" + "10012",
                "2022",
                "kystnært",
                "010012",
                "Hellefisk",
                6,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "06" + "10013",
                "2022",
                "indhandling",
                "010013",
                "Hellefisk",
                6,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "07" + "10027",
                "2022",
                "havgående",
                "010027",
                "Torsk",
                7,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "07" + "10028",
                "2022",
                "indhandling",
                "010028",
                "Torsk",
                7,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "08" + "10031",
                "2022",
                "havgående",
                "010031",
                "Kuller",
                8,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "08" + "10032",
                "2022",
                "indhandling",
                "010032",
                "Kuller",
                8,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "09" + "10033",
                "2022",
                "havgående",
                "010033",
                "Sej",
                9,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "09" + "10034",
                "2022",
                "indhandling",
                "010034",
                "Sej",
                9,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "10" + "10029",
                "2022",
                "havgående",
                "010029",
                "Rødfisk",
                10,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "10" + "10030",
                "2022",
                "indhandling",
                "010030",
                "Rødfisk",
                10,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "11" + "10011",
                "2022",
                "havgående",
                "010011",
                "Reje - havgående licens",
                11,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "12" + "10012",
                "2022",
                "kystnært",
                "010012",
                "Reje - kystnær licens",
                12,
                "Innaarsuit",
                10769,
                None,
            ],
            [
                "22" + "010769" + "13" + "10014",
                "2022",
                "svalbard",
                "010014",
                "Reje - Svalbard og Barentshavet",
                13,
                "Innaarsuit",
                10769,
                None,
            ],
        ]

        self.maxDiff = None
        actual = [[cell.value for cell in ws[row]] for row in range(1, ws.max_row + 1)]
        for item in expected:
            self.assertIn(item, actual)
